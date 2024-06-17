from flask import Flask, request, jsonify
import io
import os
import cx_Oracle
from google.cloud import vision
from google.cloud.vision_v1 import types
from PIL import Image
import openpyxl
import base64

app = Flask(__name__)

def excel_checking(excel_path):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    # Get the maximum row index
    max_row = sheet.max_row

    # Define the pairs of target data
    target_data_pairs = [
        ("ഗ്രൂപ്പ്/സെന്റർ അംഗങ്ങളുടെ പേര് &ഒപ്പ്/വിരലടയാളം", "ഡിമാൻഡ് പ്രോമിസറി നോട്ട്"),
        ("Name & Signature/Thumb Impression of Group/Centre Members", "DEMAND PROMISSORY NOTE"),
        # Add more pairs as needed
    ]

    # Iterate over each pair of target data
    for start_data, end_data in target_data_pairs:
        # Initialize variables to track rows
        start_row = None
        end_row = None

        # Iterate over each row in the worksheet
        for row in range(1, max_row + 1):
            # Check if the start target data is found in column A
            if sheet.cell(row=row, column=2).value == start_data:
                start_row = row
            # Check if the end target data is found in column A
            elif sheet.cell(row=row, column=2).value == end_data:
                end_row = row
                break  # Break the loop if the end target data is found

        # Check if start and end rows are found
        if start_row is not None and end_row is not None:
            if end_row - start_row > 7:
                return "document ok"
            elif end_row - start_row == 1:
                return "document not ok"
            else:
                return "Check the document"
        else:
            continue


# Set Google Cloud credentials
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = r"vision-377709-7ecf57251a35.json"

# Define source and destination folders
source_folder = r'F:\\my code\\vision api\\input\\'
dest_folder = r'F:\\my code\\vision api\\realinput\\'
output_folder = r"F:\\my code\\vision api\\output\\"
unmasked = r'F:\\my code\\vision api\\unmasked\\'

# Connect to Oracle database
conn = cx_Oracle.connect("mana0809", "mana0809", "MAFILUAT")
cursor = conn.cursor()

# Instantiate Google Cloud Vision client
client = vision.ImageAnnotatorClient()

# Create a new Excel workbook
wb = openpyxl.Workbook()
sheet = wb.active

# Add headers to the Excel sheet
sheet["A1"] = "Filename"
sheet["B1"] = "Extracted Text"

# Start writing from row 2
row_index = 2

folder_path='Folder which contains image'

# Process each image in the source folder
for filename in os.listdir(source_folder):
    if filename.endswith('.jpg') or filename.endswith('.tif') or filename.endswith('.bmp') or filename.endswith('.png'):
            image = Image.open(os.path.join(source_folder, filename))
            image = image.convert('RGB')
            new_filename = os.path.splitext(filename)[0] + '.jpg'
            # Perform OCR on the image with Malayalam language hint
            try:
                # Perform OCR on the image with Malayalam language hint
                response = client.text_detection(image=image)
                
                # Extract text from response
                texts = response.text_annotations
                extracted_text = texts[0].description if texts else ""  # Extracted text or empty string if no text found
                
                # Split the extracted text into lines
                lines = extracted_text.split('\n')

                # Write each line of extracted text to a new row in the Excel sheet
                for line in lines:
                    sheet.cell(row=row_index, column=1, value=filename)
                    sheet.cell(row=row_index, column=2, value=line)
                    row_index += 1
            except Exception as e:
                print("Error processing image:", str(e))
                continue

        # Save the Excel workbook
            excel_file_path = "extracted_text.xlsx"
            wb.save(excel_file_path)

            document_status = excel_checking(excel_file_path)

            print(document_status)


